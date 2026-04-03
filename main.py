import openpyxl


# Function to open and print the content of an Excel file
def print_excel_content(file_path, sheet_name):
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path)
        
        # Select the specific sheet by name
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            print(f"Error: The sheet '{sheet_name}' does not exist in the workbook.")
            return
        
        # Iterate through rows and print each cell's value
        for row in sheet.iter_rows(values_only=True):
            print(row)
    
    except FileNotFoundError:
        print(f"Error: The file '{file_path}' was not found.")
    except Exception as e:
        print(f"An error occurred: {e}")

# Specify the path to your Excel file
excel_file_path = "../Chart_of_Accounts.xlsx"  # Replace with the actual file path

# Call the function
print_excel_content(excel_file_path, sheet_name="Chart of Accounts")  # Replace "Sheet1" with the actual sheet name if needed

