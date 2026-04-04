from __future__ import annotations

import argparse
import os
from pathlib import Path

import pandas as pd
import pyodbc
from dotenv import load_dotenv


def get_connection() -> pyodbc.Connection:
    load_dotenv()

    required = ["DRIVER", "SERVER", "DATABASE", "ENTRA_USER", "ENTRA_PASS"]
    missing = [k for k in required if not os.getenv(k)]
    if missing:
        raise RuntimeError(f"Missing environment variables: {', '.join(missing)}")

    raw_driver = os.getenv("DRIVER", "").strip()
    # Accept either DRIVER=ODBC Driver 18 for SQL Server or DRIVER={ODBC Driver 18 for SQL Server}
    driver = raw_driver.strip("{}\"'")

    conn_str = (
        f"Driver={{{driver}}};"
        f"Server=tcp:{os.getenv('SERVER')},1433;"
        f"Database={os.getenv('DATABASE')};"
        f"UID={os.getenv('ENTRA_USER')};"
        f"PWD={os.getenv('ENTRA_PASS')};"
        "Encrypt=yes;"
        "TrustServerCertificate=no;"
        "Connection Timeout=60;"
        "Authentication=ActiveDirectoryPassword;"
    )
    return pyodbc.connect(conn_str)


def apply_total_formulas(output_path: Path, df: pd.DataFrame) -> None:
    """
    Add SUM formulas for total accounts:
      1999 (Assets), 2999 (Liabilities), 3999 (Equity), 4999 (Income), 5999 (Expense)
    across Opening/Debit/Credit/Closing columns.
    """
    from openpyxl import load_workbook

    # column letters in exported sheet
    col_map = {
        "Opening Balance": "C",
        "Debit": "D",
        "Credit": "E",
        "Closing Balance": "F",
    }

    # total row -> detail account range (inclusive)
    total_ranges = {
        1999: (1000, 1998),
        2999: (2000, 2998),
        3999: (3000, 3998),
        4999: (4000, 4998),
        5999: (5000, 5998),
    }

    # Map account code to Excel row number (header is row 1)
    account_to_row: dict[int, int] = {}
    for idx, account_code in enumerate(df["Account Number"].tolist(), start=2):
        try:
            account_to_row[int(account_code)] = idx
        except Exception:
            continue

    wb = load_workbook(output_path)
    ws = wb["Trial Balance"]

    for total_code, (start_code, end_code) in total_ranges.items():
        total_row = account_to_row.get(total_code)
        if total_row is None:
            continue

        detail_rows = [
            row_num
            for code, row_num in account_to_row.items()
            if start_code <= code <= end_code
        ]
        if not detail_rows:
            continue

        first_row = min(detail_rows)
        last_row = max(detail_rows)

        for col_letter in col_map.values():
            ws[f"{col_letter}{total_row}"] = f"=SUM({col_letter}{first_row}:{col_letter}{last_row})"

    # Standardize column widths
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16

    # Accounting format, 2 decimals, Symbol: None
    # Equivalent custom number format used by Excel accounting with no currency symbol.
    accounting_no_symbol_format = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
    for row in range(2, ws.max_row + 1):
        for col_letter in ("C", "D", "E", "F"):
            ws[f"{col_letter}{row}"].number_format = accounting_no_symbol_format

    wb.save(output_path)


def build_trial_balance_df(conn: pyodbc.Connection, year: int, month: int) -> pd.DataFrame:
    period_id = year * 100 + month

    sql = """
SELECT
    a.account_code AS [Account Number],
    a.account_name AS [Account Name],
    SUM(CASE WHEN je.period_id = ? AND jel.dr_cr = 'D' THEN jel.amount ELSE 0 END) AS debit_total,
    SUM(CASE WHEN je.period_id = ? AND jel.dr_cr = 'C' THEN jel.amount ELSE 0 END) AS credit_total
FROM acct.dim_account a
LEFT JOIN acct.journal_entry_line jel
    ON jel.account_id = a.account_id
LEFT JOIN acct.journal_entry je
    ON je.journal_entry_id = jel.journal_entry_id
GROUP BY
    a.account_code,
    a.account_name
ORDER BY
    a.account_code;
"""

    base_df = pd.read_sql(sql, conn, params=[period_id, period_id])

    out = pd.DataFrame()
    out["Account Number"] = base_df["Account Number"]
    out["Account Name"] = base_df["Account Name"]

    # Leave opening/closing blank for now as requested.
    out["Opening Balance"] = ""

    debit_vals = base_df["debit_total"].fillna(0)
    credit_vals = base_df["credit_total"].fillna(0)

    # If no activity for account in period, keep blank.
    out["Debit"] = [float(v) if float(v) != 0 else "" for v in debit_vals]
    out["Credit"] = [float(v) if float(v) != 0 else "" for v in credit_vals]

    out["Closing Balance"] = ""

    return out


def default_output_path(year: int, month: int) -> Path:
    # Script lives in ROC/. Output one folder up to match your existing files.
    root = Path(__file__).resolve().parent.parent
    return root / f"Trial_Balance_{month:02d}_{year}.xlsx"


def export_trial_balance(year: int, month: int, output_path: Path) -> None:
    if month < 1 or month > 12:
        raise ValueError("Month must be between 1 and 12")

    conn = get_connection()
    try:
        df = build_trial_balance_df(conn, year, month)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Trial Balance")

        apply_total_formulas(output_path, df)

        print(f"✅ Trial balance exported: {output_path}")
        print(f"Rows: {len(df)} | Period: {month:02d}/{year}")
    finally:
        conn.close()


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Create Trial_Balance_MM_YYYY.xlsx from SQL data for a selected month."
    )
    parser.add_argument("--year", type=int, default=2026, help="Year, e.g. 2026")
    parser.add_argument("--month", type=int, default=3, help="Month 1-12, e.g. 3 for March")
    parser.add_argument("--output", type=Path, default=None, help="Optional output file path")

    args = parser.parse_args()

    output = args.output if args.output else default_output_path(args.year, args.month)
    export_trial_balance(args.year, args.month, output)


if __name__ == "__main__":
    main()
